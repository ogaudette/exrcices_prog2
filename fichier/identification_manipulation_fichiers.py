# Chapitre 3.2

import json

# Écriture dans fichier JSON
data = {"nom": "olivier",
        "Sport prefere": "Soccer",
        "age": 18}
with open("data_olivier.json", "w") as f:
    json.dump(data, f, indent=4)

# Lecture dans ficier JSON
with open("data_olivier.json", "r") as f:
    data_lue = json.load(f)
    print(data_lue["nom"])

import csv

# écriture dans fichier CSV
with open("data_olivier.csv", "w", newline="") as f:
    writer = csv.writer(f)
    writer.writerow(["Nom", "Sport Prefere", "Age"])
    writer.writerow(["olivier", "Soccer", "18"])

# Lecture d'un fichier CSV
with open("data_olivier.csv", "r") as f:
    reader = csv.reader(f)
    for ligne in reader:
        print(ligne)

import xml.etree.ElementTree as ET

# création fichier XML
root = ET.Element("personne")
nom = ET.SubElement(root, "nom")
nom.text = "olivier"
tree = ET.ElementTree(root)
tree.write("data_olivier.xml")

# Lecture fichier XML
tree = ET.parse("data_olivier.xml")
root = tree.getroot()
print(root.find("nom").text)  # afficher olivier

import pickle

# sérialisation
data1 = {"clé": [1,2,3]}
with open("data1.pkl", "wb") as f:
    pickle.dump(data1, f)

# Désérialisation
with open("data1.pkl", "rb") as f:
    data1_lue = pickle.load(f)
    print(data1_lue["clé"]) # affiche 1,2,3]

from PIL import Image

# Ouverture d'une image
image = Image.open("skibidi.jpg")

# Afficher des infos
print(image.format, image.size)

# Redimensionner et sauvegarder
image_redim = image.resize((256, 256))
image_redim.save("image_redim.png","PNG")

from PyPDF2 import PdfReader

# Lecture d'un PDF
reader = PdfReader("pdf_descartes.pdf")
print(f"Nombre de pages: {len(reader.pages)}")

# extraction du texte d'une page
page = reader.pages[0]
texte = page.extract_text()
print(texte)

import pdfplumber

with pdfplumber.open("pdf_descartes.pdf") as pdf:
    page = pdf.pages[0]
    texte = page.extract_text()
    print(texte)

from openpyxl import Workbook, load_workbook

# Création d'un fichier excel
wb = Workbook()
ws = wb.active
ws["A1"] = "Nom"
ws["B1"] = "Sport Prefere"
ws["C1"] = "Age"
ws["A2"] = "olivier"
ws["B2"] = "Soccer"
ws["C2"] = 18
wb.save("data_olivier.xlsx")

# Lecture d'un fichier excel
wb = load_workbook("data_olivier.xlsx")
ws = wb.active
print(ws["A2"].value)

import pandas as pd

# Lecture d'un fichier excel
df = pd.read_excel("data_olivier.xlsx")
print(df)

# écriture dans un fichier excel
df.to_excel("nouveau.xlsx", index=False)

import mimetypes

fichier = "image.jpg"
type_mime, _ = mimetypes.guess_type(fichier)
print(type_mime)  # affiche "image/jpeg"

import magic

mime = magic.Magic(mime=True)
type_fichier = mime.from_file("image.jpg")
print(type_fichier)  # affiche "image/jpeg"
