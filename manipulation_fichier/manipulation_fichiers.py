# Exercice d'application

import os
from pathlib import Path
import json
import csv
import pickle
from PIL import Image
from PyPDF2 import PdfReader
from openpyxl import Workbook, load_workbook
import mimetypes

def exercice_1():
    try:
        chemin = input("Entrez le chemin du dossier : ")
        if not os.path.exists(chemin):
            raise FileNotFoundError("Le chemin n'existe pas")

        nb_fichier = 0
        nb_dossier = 0
        for element in os.listdir(chemin):
            chemin_complet = os.path.join(chemin, element)
            if os.path.isfile(chemin_complet):
                print(f"- {element} (fichier)")
                nb_fichier += 1
            elif os.path.isdir(chemin_complet):
                print(f"- {element} (dossier)")
                nb_dossier += 1
            else:
                print(f"- {element} (inconnu)")

        print(f"Total: {nb_fichier} fichier(s), {nb_dossier} dossier(s)")
    except FileNotFoundError as e:
        print(f"Erreur: {e}")

    except PermissionError:
        print(f"Permission refusée pour accéder au dossier")

    except Exception as e:
        print(f"Erreur inattendue survenue: {e}")

# exercice_1()

def exercice_2():  # exercice typique d'examen
    try:
        chemin1 = input("Entrez le chemin du dossier : ")
        if not Path(chemin1).exists():
            raise FileNotFoundError("Le chemin n'existe pas")
        extension = input("Entrez le extension : ")
        if not extension.startswith("."):
            extension = "." + extension
        fichiers = Path(chemin1).rglob(f"*{extension}")
        for fichier in fichiers:
            print(f"- {fichier}")

    except FileNotFoundError as e:
        print(f"Erreur:{e}")
    except Exception as e:
        print(f"Erreur inattendue est survenue : {e}")

# exercice_2()

def exercice_3():
    try:
        etudiant = {"nom":"Olivier", "age":18, "notes":[15,18,12]}
        with open("etudiant.json", "w") as f:
            json.dump(etudiant, f, indent=4)

        with open("etudiant.json", "r") as f:
            data = json.load(f)
            print(f"Les notes de l'étudiant son: {data['notes']}")

    except IOError:
        print("Problème lors accès fichier json")
    except Exception as e:
        print(f"Erreur inattendue: {e}")

# exercice_3()

def exercice_4():
    try:
        with open("etudiant.csv", "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["Nom", "Math", "Info"])  # en-tête
            writer.writerow(["Alice", 15, 18])
            writer.writerow(["oli", 13, 19])
            writer.writerow(["Bob", 12, 16])
        math = 0
        info = 0
        nb_fois = 0
        with open("etudiant.csv", "r") as f:
            reader = csv.reader(f)
            next(reader)  # ignore l'en-tête
            for ligne in reader:
                math += float(ligne[1])
                info += float(ligne[2])
                nb_fois += 1

            moyenne_math = math / nb_fois
            moyenne_info = info / nb_fois

            print(f"Moyenne Maths: {moyenne_math:.1f}")
            print(f"Moyenne Info: {moyenne_info:.1f}")

    except FileNotFoundError as e:
        print(f"fichier pas trouvé: {e}")
    except Exception as e:
        print(f"Erreur inattendue: {e}")

# exercice_4()
def exercice_5():
    nombres = [1, 3, 5, 7, 9]
    with open("nombres.pkl", "wb") as f:
        pickle.dump(nombres, f)

    with open("nombres.pkl", "rb") as f:
        nombres.lue = pickle.load(f)
        print(sum(nombres))

exercice_5()

def exercice_6():
    try:
        image = Image.open("sango.jpg")
        print(f"Format : {image.format}")
        print(f"Taille : {image.size}")

        largeur, hauteur = image.size
        nouvelle_taille = (largeur // 2, hauteur // 2)
        image_redim = image.resize(nouvelle_taille)
        image_redim.save("photo_redim.jpg")
        print(f"Nouvelle taille : {image_redim.size}")
    except FileNotFoundError:
        print("Erreur : L’image 'photo.jpg' n’existe pas.")
    except Exception as e:
        print(f"Une erreur inattendue est survenue : {e}")

def exercice_7():
    try:
        reader = PdfReader("texte.pdf")
        nb_pages = len(reader.pages)
        print(f"Nombre de pages : {nb_pages}")

        page = reader.pages[0]
        texte = page.extract_text()
        print(f"Texte de la page 1 : {texte}")
    except FileNotFoundError:
        print("Erreur : Le fichier PDF 'texte.pdf' n’existe pas.")
    except Exception as e:
        print(f"Une erreur inattendue est survenue : {e}")

def exercice_8():
    try:
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Nom"
        ws["B1"] = "Note"
        ws["A2"] = "Alice"
        ws["B2"] = 15
        ws["A3"] = "Bob"
        ws["B3"] = 12
        ws["A4"] = "Charlie"
        ws["B4"] = 18
        ws["A5"] = "David"
        ws["B5"] = 14
        wb.save("resultats.xlsx")

        wb = load_workbook("resultats.xlsx")
        ws = wb.active
        meilleure_note = 0
        meilleur_etudiant = ""
        for row in ws.iter_rows(min_row=2, values_only=True):
            nom, note = row
            if note > meilleure_note:
                meilleure_note = note
                meilleur_etudiant = nom
        print(f"Meilleure note : {meilleur_etudiant} ({meilleure_note})")
    except IOError:
        print("Erreur : Problème lors de l’accès au fichier Excel.")
    except Exception as e:
        print(f"Une erreur inattendue est survenue : {e}")

def exercice_9():
    try:
        fichier = input("Entrez le nom du fichier : ")
        type_mime, _ = mimetypes.guess_type(fichier)
        print(f"Type MIME : {type_mime}")

        if type_mime:
            if "image" in type_mime:
                print("C’est une image")
            elif "pdf" in type_mime:
                print("C’est un document PDF")
            elif "text" in type_mime:
                print("C’est un fichier texte")
            else:
                print("Type inconnu")
        else:
            print("Impossible de déterminer le type")
    except Exception as e:
        print(f"Une erreur inattendue est survenue : {e}")
